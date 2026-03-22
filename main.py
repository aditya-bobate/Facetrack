# ================= FACE TRACK — REFINED UI =================
# Applies the polished design system to the FULL REPLACE VERSION
# Features: Login, Register (student/admin), Admin Dash, Head Admin Dash
# Head Admin: DELETE students + logout

import tkinter as tk
from tkinter import messagebox
import sqlite3
from datetime import datetime
import cv2, os, time, pickle, shutil
from openpyxl import Workbook, load_workbook

# ══════════════════════════════════════════════════════════
#  DESIGN TOKENS
# ══════════════════════════════════════════════════════════

C = {
    "bg":          "#0A0E14",
    "surface":     "#111720",
    "card":        "#151D2A",
    "card2":       "#1A2435",
    "border":      "#1E2D42",
    "border_hi":   "#2A4A6B",
    "accent":      "#4A9EFF",
    "accent2":     "#2979CC",
    "accent_glow": "#1A3A5C",
    "green":       "#34D399",
    "green_dim":   "#0D2E1E",
    "amber":       "#FBBF24",
    "amber_dim":   "#2E1F05",
    "red":         "#F87171",
    "red_dim":     "#2E0D0D",
    "red_hi":      "#5C1A1A",
    "purple":      "#A78BFA",
    "purple_dim":  "#1E1040",
    "t1":          "#EFF6FF",
    "t2":          "#94A3B8",
    "t3":          "#4A5568",
    "t4":          "#2D3748",
}

F = {
    "logo":  ("Helvetica", 20, "bold"),
    "h1":    ("Helvetica", 18, "bold"),
    "h2":    ("Helvetica", 14, "bold"),
    "h3":    ("Helvetica", 11, "bold"),
    "body":  ("Helvetica", 11),
    "small": ("Helvetica", 9),
    "micro": ("Helvetica", 8),
    "input": ("Helvetica", 11),
    "btn":   ("Helvetica", 11, "bold"),
    "stat":  ("Helvetica", 24, "bold"),
}

# ══════════════════════════════════════════════════════════
#  DATABASE
# ══════════════════════════════════════════════════════════

conn = sqlite3.connect("facetrack.db")
cur  = conn.cursor()

for sql in [
    """CREATE TABLE IF NOT EXISTS users(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE, password TEXT,
        role TEXT, enrollment TEXT)""",
    """CREATE TABLE IF NOT EXISTS subjects(
        id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT UNIQUE)""",
    """CREATE TABLE IF NOT EXISTS attendance(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT, enrollment TEXT, subject TEXT,
        date TEXT, in_time TEXT, status TEXT)""",
]:
    cur.execute(sql)
conn.commit()

cur.execute("INSERT OR IGNORE INTO users(username,password,role) VALUES (?,?,?)",
            ("headadmin","admin123","head_admin"))
conn.commit()

# ══════════════════════════════════════════════════════════
#  FACE MODEL
# ══════════════════════════════════════════════════════════

model = cv2.face.LBPHFaceRecognizer_create()
if os.path.exists("models/lbph_model.xml"):
    model.read("models/lbph_model.xml")

label_map = {}
if os.path.exists("models/labels.pkl"):
    with open("models/labels.pkl","rb") as f:
        label_map = pickle.load(f)

face_cascade = cv2.CascadeClassifier("haarcascade_frontalface_default.xml")

# ══════════════════════════════════════════════════════════
#  ROOT WINDOW
# ══════════════════════════════════════════════════════════

root = tk.Tk()
root.geometry("500x680")
root.title("FaceTrack")
root.configure(bg=C["bg"])
root.resizable(False, False)

current_user  = None
lecture_start = None
LATE_LIMIT    = 900

# ══════════════════════════════════════════════════════════
#  ANIMATION ENGINE
# ══════════════════════════════════════════════════════════

_anim_jobs = []

def lerp_color(c1, c2, t):
    r1,g1,b1 = int(c1[1:3],16), int(c1[3:5],16), int(c1[5:7],16)
    r2,g2,b2 = int(c2[1:3],16), int(c2[3:5],16), int(c2[5:7],16)
    r = max(0, min(255, int(r1+(r2-r1)*t)))
    g = max(0, min(255, int(g1+(g2-g1)*t)))
    b = max(0, min(255, int(b1+(b2-b1)*t)))
    return f"#{r:02x}{g:02x}{b:02x}"

def ease_out(t):
    return 1-(1-t)**3

def animate_btn(btn, from_bg, to_bg, from_fg, to_fg, steps=10, delay=12, _step=0):
    if _step > steps: return
    t = ease_out(_step/steps)
    try:
        btn.configure(bg=lerp_color(from_bg,to_bg,t),
                      fg=lerp_color(from_fg,to_fg,t))
    except tk.TclError:
        return
    job = root.after(delay, animate_btn, btn, from_bg, to_bg, from_fg, to_fg,
                     steps, delay, _step+1)
    _anim_jobs.append(job)

def pulse_logo(lbl, colors, idx=0):
    try:
        lbl.configure(fg=colors[idx % len(colors)])
    except tk.TclError:
        return
    job = root.after(1400, pulse_logo, lbl, colors, idx+1)
    _anim_jobs.append(job)

def cancel_anims():
    for j in _anim_jobs:
        try: root.after_cancel(j)
        except: pass
    _anim_jobs.clear()

# ══════════════════════════════════════════════════════════
#  SCROLLABLE BODY
# ══════════════════════════════════════════════════════════

def make_scrollable_body():
    outer = tk.Frame(root, bg=C["bg"])
    outer.pack(fill="both", expand=True)

    canvas = tk.Canvas(outer, bg=C["bg"], bd=0, highlightthickness=0)
    canvas.pack(side="left", fill="both", expand=True)

    sb = tk.Scrollbar(outer, orient="vertical", command=canvas.yview,
                      troughcolor=C["bg"], bg=C["border"],
                      activebackground=C["accent2"], width=5,
                      relief="flat", bd=0)
    canvas.configure(yscrollcommand=sb.set)

    frame = tk.Frame(canvas, bg=C["bg"])
    win_id = canvas.create_window((0,0), window=frame, anchor="nw")

    def on_configure(e):
        canvas.configure(scrollregion=canvas.bbox("all"))
        canvas.itemconfig(win_id, width=canvas.winfo_width())
        if frame.winfo_reqheight() > canvas.winfo_height():
            sb.pack(side="right", fill="y")
        else:
            sb.pack_forget()

    frame.bind("<Configure>", on_configure)
    canvas.bind("<Configure>", on_configure)

    def _wheel(e): canvas.yview_scroll(int(-1*(e.delta/120)), "units")
    canvas.bind_all("<MouseWheel>", _wheel)

    return frame

# ══════════════════════════════════════════════════════════
#  COMPONENT LIBRARY
# ══════════════════════════════════════════════════════════

def clear():
    cancel_anims()
    for w in root.winfo_children(): w.destroy()

def get_subjects():
    return [r[0] for r in cur.execute("SELECT name FROM subjects").fetchall()]

def Card(parent, padx=20, pady=16, gap=0):
    outer = tk.Frame(parent, bg=C["border"], padx=1, pady=1)
    outer.pack(fill="x", padx=24, pady=(gap, 0))
    inner = tk.Frame(outer, bg=C["card"])
    inner.pack(fill="x")
    body  = tk.Frame(inner, bg=C["card"], padx=padx, pady=pady)
    body.pack(fill="x")
    return body

def Lbl(p, text, font=None, color=None, bg=None, anchor="w", pady=0):
    return tk.Label(p, text=text, font=font or F["body"],
                    fg=color or C["t1"], bg=bg or C["card"],
                    anchor=anchor, pady=pady)

def Divider(parent, color=None, pady=10):
    tk.Frame(parent, height=1, bg=color or C["border"]).pack(fill="x", pady=pady)

# ── FANCY INPUT WITH PLACEHOLDER ─────────────────────────
class FancyEntry:
    def __init__(self, parent, placeholder="", show=None, bg=None):
        self.placeholder = placeholder
        self.show_char   = show
        self.showing_ph  = True
        self._bg         = bg or C["card2"]

        self.frame = tk.Frame(parent, bg=C["border"], padx=1, pady=1)
        self.inner = tk.Frame(self.frame, bg=self._bg)
        self.inner.pack(fill="x")

        self.entry = tk.Entry(
            self.inner, font=F["input"],
            bg=self._bg, fg=C["t3"],
            insertbackground=C["accent"],
            relief="flat", bd=10,
            selectbackground=C["accent_glow"],
            selectforeground=C["t1"],
        )
        self.entry.pack(fill="x")
        self.entry.insert(0, placeholder)
        self.entry.bind("<FocusIn>",  self._focus_in)
        self.entry.bind("<FocusOut>", self._focus_out)

    def _focus_in(self, e):
        self.frame.configure(bg=C["accent"])
        if self.showing_ph:
            self.entry.delete(0, "end")
            self.entry.configure(fg=C["t1"], show=self.show_char or "")
            self.showing_ph = False

    def _focus_out(self, e):
        self.frame.configure(bg=C["border"])
        if not self.entry.get():
            self.entry.configure(fg=C["t3"], show="")
            self.entry.insert(0, self.placeholder)
            self.showing_ph = True

    def get(self):
        return "" if self.showing_ph else self.entry.get()

    def pack(self, **kw):
        self.frame.pack(**kw); return self

# ── ANIMATED BUTTON ───────────────────────────────────────
class AnimBtn:
    STYLES = {
        "primary": (C["accent_glow"], C["accent"],  C["accent"],  C["bg"]),
        "success": (C["green_dim"],   C["green"],   C["green"],   C["bg"]),
        "warning": (C["amber_dim"],   C["amber"],   C["amber"],   C["bg"]),
        "danger":  (C["red_dim"],     C["red"],     C["red"],     C["bg"]),
        "purple":  (C["purple_dim"],  C["purple"],  C["purple"],  C["bg"]),
        "ghost":   (C["card"],        C["t3"],      C["card2"],   C["t2"]),
    }

    def __init__(self, parent, text, command, style="primary", icon="", full_width=False):
        nb, nfg, hb, hfg = self.STYLES.get(style, self.STYLES["primary"])
        self.nb=nb; self.nfg=nfg; self.hb=hb; self.hfg=hfg
        label = f"{icon}  {text}" if icon else text

        self.frame = tk.Frame(parent, bg=nfg, padx=1, pady=1)
        self.btn = tk.Button(
            self.frame, text=label, command=command,
            font=F["btn"], fg=nfg, bg=nb,
            activeforeground=hfg, activebackground=hb,
            relief="flat", bd=0, cursor="hand2",
            padx=16, pady=10,
        )
        self.btn.pack(fill="x" if full_width else "none")
        self.btn.bind("<Enter>", lambda e: animate_btn(self.btn, nb, hb, nfg, hfg))
        self.btn.bind("<Leave>", lambda e: animate_btn(self.btn, hb, nb, hfg, nfg))
        self.btn.bind("<ButtonPress-1>",   lambda e: self.btn.configure(pady=8,  padx=18))
        self.btn.bind("<ButtonRelease-1>", lambda e: self.btn.configure(pady=10, padx=16))

    def pack(self, **kw):
        self.frame.pack(**kw); return self

# ── RADIO CHIP ────────────────────────────────────────────
class RadioChip:
    def __init__(self, parent, text, variable, value, on_col, on_dim, callback=None):
        self.var=variable; self.val=value
        self.on_col=on_col; self.on_dim=on_dim; self.cb=callback

        self.frame = tk.Frame(parent, bg=C["border"], padx=1, pady=1, cursor="hand2")
        self.inner = tk.Frame(self.frame, bg=C["card2"])
        self.inner.pack(fill="both")
        self.lbl = tk.Label(self.inner, text=text, font=F["h3"],
                            fg=C["t3"], bg=C["card2"], padx=14, pady=9)
        self.lbl.pack()
        for w in (self.frame, self.inner, self.lbl):
            w.bind("<Button-1>", self._click)
        self._refresh()

    def _click(self, e=None):
        self.var.set(self.val)
        if self.cb: self.cb()

    def _refresh(self):
        sel = self.var.get() == self.val
        if sel:
            self.frame.configure(bg=self.on_col)
            self.inner.configure(bg=self.on_dim)
            self.lbl.configure(fg=self.on_col, bg=self.on_dim)
        else:
            self.frame.configure(bg=C["border"])
            self.inner.configure(bg=C["card2"])
            self.lbl.configure(fg=C["t3"], bg=C["card2"])

    def pack(self, **kw):
        self.frame.pack(**kw); return self

def StatusBar(text, variant="muted"):
    col = {"muted": C["t4"], "success": C["green"],
           "warning": C["amber"], "info": C["accent"], "danger": C["red"]}
    bar = tk.Frame(root, bg=C["surface"])
    bar.pack(side="bottom", fill="x")
    tk.Frame(bar, height=1, bg=C["border"]).pack(fill="x")
    row = tk.Frame(bar, bg=C["surface"])
    row.pack(fill="x", padx=16, pady=5)
    tk.Label(row, text="●", font=F["micro"],
             fg=col.get(variant, C["t4"]), bg=C["surface"]).pack(side="left", padx=(0,5))
    tk.Label(row, text=text, font=F["micro"], fg=C["t3"], bg=C["surface"]).pack(side="left")

def TopNav(subtitle="", badge="", badge_color=C["accent"], back_cmd=None):
    nav = tk.Frame(root, bg=C["surface"])
    nav.pack(fill="x")
    tk.Frame(nav, height=2, bg=C["accent"]).pack(fill="x")

    row = tk.Frame(nav, bg=C["surface"])
    row.pack(fill="x", padx=20, pady=12)

    if back_cmd:
        b = tk.Button(row, text="‹", font=("Helvetica",16,"bold"),
                      fg=C["t2"], bg=C["surface"],
                      activeforeground=C["accent"], activebackground=C["surface"],
                      relief="flat", bd=0, cursor="hand2", command=back_cmd)
        b.pack(side="left", padx=(0,10))
        b.bind("<Enter>", lambda e: b.configure(fg=C["accent"]))
        b.bind("<Leave>", lambda e: b.configure(fg=C["t2"]))

    logo = tk.Label(row, text="⬡ FaceTrack", font=F["logo"],
                    fg=C["accent"], bg=C["surface"])
    logo.pack(side="left")
    pulse_logo(logo, [C["accent"], C["accent2"], "#5AB4FF", C["accent"]])

    if badge:
        tk.Label(row, text=f" {badge} ", font=F["micro"],
                 fg=badge_color, bg=C["bg"], padx=6, pady=3).pack(side="left", padx=10)

    if subtitle:
        tk.Label(row, text=subtitle, font=F["small"],
                 fg=C["t3"], bg=C["surface"]).pack(side="left", padx=10)

    tk.Frame(nav, height=1, bg=C["border"]).pack(fill="x")

# ══════════════════════════════════════════════════════════
#  CORE FUNCTIONS  (original logic — unchanged)
# ══════════════════════════════════════════════════════════

def capture_face(username):
    os.makedirs(f"dataset/{username}", exist_ok=True)
    cam=cv2.VideoCapture(0); count=0; last=time.time()
    while True:
        ret,frame=cam.read()
        gray=cv2.cvtColor(frame,cv2.COLOR_BGR2GRAY)
        faces=face_cascade.detectMultiScale(gray,1.3,5)
        for (x,y,w,h) in faces:
            face=cv2.resize(gray[y:y+h,x:x+w],(200,200))
            if time.time()-last>2 and count<20:
                count+=1; cv2.imwrite(f"dataset/{username}/{count}.jpg",face); last=time.time()
            cv2.rectangle(frame,(x,y),(x+w,y+h),(74,158,255),2)
        cv2.putText(frame,f"Captured: {count}/20",(10,30),
                    cv2.FONT_HERSHEY_SIMPLEX,1,(74,158,255),2)
        cv2.imshow("FaceTrack — Face Registration",frame)
        if count>=20 or cv2.waitKey(1)==27: break
    cam.release(); cv2.destroyAllWindows()

def start_attendance(subject):
    global lecture_start
    lecture_start=time.time()
    cam=cv2.VideoCapture(0); marked=set(); timer={}
    while True:
        ret,frame=cam.read()
        if not ret: break
        gray=cv2.cvtColor(frame,cv2.COLOR_BGR2GRAY)
        faces=face_cascade.detectMultiScale(gray,1.3,5)
        for (x,y,w,h) in faces:
            face=cv2.resize(gray[y:y+h,x:x+w],(200,200))
            label,conf=model.predict(face)
            u=label_map.get(label)
            col=(52,211,153) if conf<75 else (248,113,113)
            cv2.rectangle(frame,(x,y),(x+w,y+h),col,2)
            cv2.putText(frame,f"{u}",(x,y-10),cv2.FONT_HERSHEY_SIMPLEX,0.65,col,2)
            if conf<75:
                timer.setdefault(u,time.time())
                if time.time()-timer[u]>2 and u not in marked:
                    marked.add(u)
                    status="On Time" if time.time()-lecture_start<=LATE_LIMIT else "Late"
                    row=cur.execute("SELECT enrollment FROM users WHERE username=?",(u,)).fetchone()
                    enr=row[0] if row else ""
                    cur.execute(
                        "INSERT INTO attendance(username,enrollment,subject,date,in_time,status)"
                        " VALUES (?,?,?,?,?,?)",
                        (u,enr,subject,datetime.now().strftime("%Y-%m-%d"),
                         datetime.now().strftime("%H:%M:%S"),status))
                    conn.commit()
        cv2.putText(frame,f"Subject: {subject}  |  Marked: {len(marked)}",(10,30),
                    cv2.FONT_HERSHEY_SIMPLEX,0.75,(74,158,255),2)
        cv2.putText(frame,"ESC to stop",(10,frame.shape[0]-15),
                    cv2.FONT_HERSHEY_SIMPLEX,0.5,(148,163,184),1)
        cv2.imshow(f"FaceTrack — {subject}",frame)
        if cv2.waitKey(1)==27: break
    cam.release(); cv2.destroyAllWindows()
    messagebox.showinfo("Session Ended",f"Marked {len(marked)} student(s) for '{subject}'.")

def export_to_excel():
    rows=cur.execute(
        "SELECT username,enrollment,subject,date,in_time,status FROM attendance"
    ).fetchall()
    if not rows: messagebox.showinfo("Export","No records found."); return
    fname="attendance.xlsx"
    wb=load_workbook(fname) if os.path.exists(fname) else Workbook()
    for u,e,s,d,t,st in rows:
        ws=wb[s] if s in wb.sheetnames else wb.create_sheet(title=s)
        if ws.max_row==1: ws.append(["Username","Enrollment","Date","Time","Status"])
        ws.append([u,e,d,t,st])
    if "Sheet" in wb.sheetnames: wb.remove(wb["Sheet"])
    wb.save(fname)
    messagebox.showinfo("Export Successful",f"Saved → {os.path.abspath(fname)}")

def delete_student(name):
    if not messagebox.askyesno(
        "Confirm Delete",
        f"Permanently delete student '{name}'?\n\n"
        f"This will remove their account, attendance records, and face data.\n"
        f"This action cannot be undone."
    ): return
    cur.execute("DELETE FROM users      WHERE username=?",(name,))
    cur.execute("DELETE FROM attendance WHERE username=?",(name,))
    conn.commit()
    if os.path.exists(f"dataset/{name}"):
        shutil.rmtree(f"dataset/{name}")
    messagebox.showinfo("Deleted",f"Student '{name}' has been permanently removed.")
    head_admin_dash()

# ══════════════════════════════════════════════════════════
#  SCREENS
# ══════════════════════════════════════════════════════════

# ── LOGIN ─────────────────────────────────────────────────
def login_screen():
    clear()
    root.configure(bg=C["bg"])
    TopNav()
    body = make_scrollable_body()

    hero = tk.Frame(body, bg=C["bg"]); hero.pack(fill="x", pady=(28,4))
    tk.Label(hero, text="Welcome back",   font=F["h1"],    fg=C["t1"], bg=C["bg"]).pack()
    tk.Label(hero, text="Sign in to access FaceTrack",
             font=F["small"], fg=C["t3"], bg=C["bg"]).pack(pady=(3,0))

    card = Card(body, gap=20)
    Lbl(card, "USERNAME", F["micro"], C["t3"]).pack(anchor="w")
    ufield = FancyEntry(card, "Enter your username")
    ufield.pack(fill="x", pady=(4,12))

    Lbl(card, "PASSWORD", F["micro"], C["t3"]).pack(anchor="w")
    pfield = FancyEntry(card, "Enter your password", show="*")
    pfield.pack(fill="x", pady=(4,16))

    def do_login():
        u=ufield.get().strip(); p=pfield.get().strip()
        r=cur.execute(
            "SELECT role FROM users WHERE username=? AND password=?",(u,p)
        ).fetchone()
        if not r:
            messagebox.showerror("Sign In Failed","Incorrect username or password."); return
        global current_user; current_user=u
        if   r[0]=="admin":        admin_dash()
        elif r[0]=="head_admin":   head_admin_dash()
        elif r[0]=="pending_admin":
            messagebox.showinfo("Pending","Your account is awaiting HOD approval.")
        else:
            messagebox.showinfo("Access","Logged in.")

    AnimBtn(card, "Sign In", do_login, "primary", "→", full_width=True).pack(fill="x")
    Divider(card, pady=14)

    foot = tk.Frame(card, bg=C["card"]); foot.pack()
    tk.Label(foot, text="New user?", font=F["small"], fg=C["t3"], bg=C["card"]).pack(side="left")
    link = tk.Label(foot, text="  Create an account", font=F["small"],
                    fg=C["accent"], bg=C["card"], cursor="hand2")
    link.pack(side="left")
    link.bind("<Button-1>", lambda e: register_screen())
    link.bind("<Enter>",    lambda e: link.configure(fg=C["accent2"]))
    link.bind("<Leave>",    lambda e: link.configure(fg=C["accent"]))

    rcard = Card(body, gap=10)
    Lbl(rcard, "ACCOUNT TYPES", F["micro"], C["t3"]).pack(anchor="w", pady=(0,10))
    for icon,role,desc,col,dim in [
        ("⬡","Student", "Face capture on registration", C["accent"], C["accent_glow"]),
        ("⬡","Admin",   "Teacher — manage attendance",  C["green"],  C["green_dim"]),
        ("⬡","HOD",     "Head admin — manage students", C["purple"], C["purple_dim"]),
    ]:
        row = tk.Frame(rcard, bg=dim, padx=12, pady=8); row.pack(fill="x", pady=3)
        tk.Label(row, text=icon,           font=F["body"],  fg=col,     bg=dim).pack(side="left")
        tk.Label(row, text=f"  {role}",    font=F["h3"],   fg=col,     bg=dim).pack(side="left")
        tk.Label(row, text=f"  —  {desc}", font=F["small"],fg=C["t3"], bg=dim).pack(side="left")

    tk.Frame(body, height=20, bg=C["bg"]).pack()
    StatusBar("FaceTrack  ·  LBPH Face Recognition Engine")


# ── REGISTER ─────────────────────────────────────────────
def register_screen():
    clear()
    root.configure(bg=C["bg"])
    TopNav(subtitle="Create Account", back_cmd=login_screen)
    body = make_scrollable_body()

    tk.Label(body, text="Create Account", font=F["h1"],
             fg=C["t1"], bg=C["bg"]).pack(pady=(20,2))
    tk.Label(body, text="Fill in your details and choose a role",
             font=F["small"], fg=C["t3"], bg=C["bg"]).pack(pady=(0,4))

    card = Card(body, gap=16)

    Lbl(card, "USERNAME",         F["micro"], C["t3"]).pack(anchor="w")
    u_f = FancyEntry(card, "Choose a username")
    u_f.pack(fill="x", pady=(4,12))

    Lbl(card, "PASSWORD",         F["micro"], C["t3"]).pack(anchor="w")
    p_f = FancyEntry(card, "Create a password", show="*")
    p_f.pack(fill="x", pady=(4,12))

    Lbl(card, "CONFIRM PASSWORD", F["micro"], C["t3"]).pack(anchor="w")
    c_f = FancyEntry(card, "Repeat your password", show="*")
    c_f.pack(fill="x", pady=(4,16))

    Divider(card, pady=8)
    Lbl(card, "ACCOUNT TYPE", F["micro"], C["t3"]).pack(anchor="w", pady=(0,8))

    role_var = tk.StringVar(value="student")
    chip_row = tk.Frame(card, bg=C["card"]); chip_row.pack(fill="x")

    enroll_lbl = Lbl(card, "ENROLLMENT NUMBER", F["micro"], C["t3"])
    e_f = FancyEntry(card, "e.g. EN2024001")
    chips = []

    def toggle_role():
        for ch in chips: ch._refresh()
        if role_var.get()=="student":
            enroll_lbl.pack(anchor="w", pady=(12,0))
            e_f.pack(fill="x", pady=(4,0))
        else:
            enroll_lbl.pack_forget(); e_f.frame.pack_forget()

    chips.append(RadioChip(chip_row,"  Student  ",role_var,"student",
                           C["accent"],C["accent_glow"],toggle_role))
    chips.append(RadioChip(chip_row,"  Admin    ",role_var,"pending_admin",
                           C["green"], C["green_dim"],  toggle_role))
    chips[0].pack(side="left", padx=(0,6)); chips[1].pack(side="left")

    enroll_lbl.pack(anchor="w", pady=(12,0))
    e_f.pack(fill="x", pady=(4,0))
    Divider(card, pady=14)

    def do_register():
        u=u_f.get().strip(); p=p_f.get(); c=c_f.get(); e=e_f.get().strip()
        if not u: messagebox.showwarning("Missing","Please enter a username."); return
        if not p: messagebox.showwarning("Missing","Please enter a password."); return
        if p!=c:  messagebox.showerror("Error","Passwords do not match."); return
        try:
            cur.execute(
                "INSERT INTO users(username,password,role,enrollment) VALUES (?,?,?,?)",
                (u,p,role_var.get(),e))
            conn.commit()
            if role_var.get()=="student": capture_face(u)
            messagebox.showinfo("Success",
                "Student registered!" if role_var.get()=="student"
                else "Admin account created.\nAwaiting HOD approval.")
            login_screen()
        except Exception:
            messagebox.showerror("Error","Username already exists. Choose another.")

    AnimBtn(card,"Create Account",do_register,"success","✓",full_width=True).pack(fill="x")
    tk.Frame(body, height=20, bg=C["bg"]).pack()
    StatusBar("Students require face capture  ·  Admin accounts need HOD approval")


# ── ADMIN (TEACHER) DASHBOARD ─────────────────────────────
def admin_dash():
    clear()
    root.configure(bg=C["bg"])
    TopNav(subtitle="Teacher Dashboard", badge="ADMIN", badge_color=C["green"])
    body = make_scrollable_body()

    gb = tk.Frame(body, bg=C["green_dim"]); gb.pack(fill="x", padx=24, pady=(16,0))
    gi = tk.Frame(gb,   bg=C["green_dim"]); gi.pack(fill="x", padx=14, pady=10)
    tk.Label(gi, text=f"●  Signed in as  {current_user}",
             font=F["h3"], fg=C["green"], bg=C["green_dim"]).pack(side="left")
    tk.Label(gi, text=datetime.now().strftime("  ·  %A, %d %b %Y"),
             font=F["small"], fg=C["t3"], bg=C["green_dim"]).pack(side="left")

    sc = Card(body, gap=12)
    Lbl(sc, "Attendance Session", F["h2"], C["t1"]).pack(anchor="w")
    Lbl(sc, "Select a subject then start the session",
        F["small"], C["t2"]).pack(anchor="w", pady=(2,0))
    Divider(sc, pady=10)

    subjects = get_subjects() or ["— No Subjects Yet —"]
    Lbl(sc, "SELECT SUBJECT", F["micro"], C["t3"]).pack(anchor="w")

    lb_frame = tk.Frame(sc, bg=C["border"], padx=1, pady=1)
    lb_frame.pack(fill="x", pady=(4,12))
    lb_inner = tk.Frame(lb_frame, bg=C["card2"]); lb_inner.pack(fill="x")
    lb = tk.Listbox(
        lb_inner, listvariable=tk.StringVar(value=subjects),
        font=F["input"], selectmode="single",
        bg=C["card2"], fg=C["t1"],
        selectbackground=C["accent_glow"], selectforeground=C["accent"],
        activestyle="none", relief="flat", bd=8, highlightthickness=0,
        height=min(len(subjects),4)
    )
    lb.pack(fill="x"); lb.selection_set(0)

    def get_sel():
        sel=lb.curselection(); return subjects[sel[0]] if sel else subjects[0]

    Divider(sc, color=C["border"], pady=6)
    Lbl(sc, "ADD NEW SUBJECT", F["micro"], C["t3"]).pack(anchor="w")

    add_row = tk.Frame(sc, bg=C["card"]); add_row.pack(fill="x", pady=(4,0))
    new_f = FancyEntry(add_row, "e.g. Mathematics", bg=C["card2"])
    new_f.frame.pack(side="left", fill="x", expand=True, padx=(0,8))

    def add_subject():
        n=new_f.get().strip()
        if not n: messagebox.showwarning("Input Required","Enter a subject name."); return
        try:
            cur.execute("INSERT INTO subjects(name) VALUES (?)",(n,))
            conn.commit(); admin_dash()
        except Exception:
            messagebox.showerror("Error","Subject already exists.")

    AnimBtn(add_row,"Add",add_subject,"primary","+").pack(side="left")

    ac = Card(body, gap=10)
    Lbl(ac, "Quick Actions", F["h3"], C["t2"]).pack(anchor="w", pady=(0,10))
    AnimBtn(ac,"Start Attendance Session",
            lambda:start_attendance(get_sel()),
            "success","◉",full_width=True).pack(fill="x", pady=(0,6))
    AnimBtn(ac,"Export Attendance to Excel",
            export_to_excel,"warning","⬇",full_width=True).pack(fill="x", pady=(0,6))
    Divider(ac, pady=8)
    AnimBtn(ac,"Sign Out",login_screen,"ghost","←",full_width=True).pack(fill="x")

    total=cur.execute("SELECT COUNT(*) FROM attendance").fetchone()[0]
    today=cur.execute("SELECT COUNT(*) FROM attendance WHERE date=?",
                      (datetime.now().strftime("%Y-%m-%d"),)).fetchone()[0]
    nsub=len(get_subjects())

    stc = Card(body, gap=8)
    Lbl(stc, "Statistics", F["micro"], C["t3"]).pack(anchor="w", pady=(0,8))
    sr = tk.Frame(stc, bg=C["card"]); sr.pack(fill="x")
    for val,label,col in [(total,"Total Records",C["accent"]),
                          (today,"Today",        C["green"]),
                          (nsub, "Subjects",     C["amber"])]:
        sf=tk.Frame(sr,bg=C["card2"],padx=1,pady=1)
        sf.pack(side="left",fill="x",expand=True,padx=(0,6))
        si=tk.Frame(sf,bg=C["card2"]); si.pack(fill="x",padx=12,pady=10)
        tk.Label(si,text=str(val), font=F["stat"],  fg=col,     bg=C["card2"]).pack()
        tk.Label(si,text=label,    font=F["micro"], fg=C["t3"], bg=C["card2"]).pack()

    tk.Frame(body, height=20, bg=C["bg"]).pack()
    StatusBar(f"Records: {total}  ·  Today: {today}  ·  Subjects: {nsub}","success")


# ── HEAD ADMIN DASHBOARD ──────────────────────────────────
def head_admin_dash():
    clear()
    root.configure(bg=C["bg"])
    TopNav(subtitle="HOD Portal", badge="HEAD ADMIN", badge_color=C["purple"])
    body = make_scrollable_body()

    gb = tk.Frame(body, bg=C["purple_dim"]); gb.pack(fill="x", padx=24, pady=(16,0))
    gi = tk.Frame(gb,   bg=C["purple_dim"]); gi.pack(fill="x", padx=14, pady=10)
    tk.Label(gi, text=f"◆  HOD Dashboard  —  {current_user}",
             font=F["h3"], fg=C["purple"], bg=C["purple_dim"]).pack(side="left")
    tk.Label(gi, text=datetime.now().strftime("  ·  %A, %d %b %Y"),
             font=F["small"], fg=C["t3"], bg=C["purple_dim"]).pack(side="left")

    # ── Student Management ────────────────────────────────
    dcard = Card(body, gap=12)
    Lbl(dcard, "Student Management", F["h2"], C["t1"]).pack(anchor="w")
    Lbl(dcard, "View and permanently remove student accounts",
        F["small"], C["t2"]).pack(anchor="w", pady=(2,0))
    Divider(dcard, pady=10)

    students = cur.execute(
        "SELECT username, enrollment FROM users WHERE role='student'"
    ).fetchall()

    if not students:
        ef = tk.Frame(dcard, bg=C["card"]); ef.pack(fill="x", pady=16)
        tk.Label(ef, text="◈", font=("Helvetica",28), fg=C["t4"], bg=C["card"]).pack()
        tk.Label(ef, text="No students registered yet",
                 font=F["body"], fg=C["t2"], bg=C["card"]).pack(pady=(4,0))
        tk.Label(ef, text="Students appear here after registration",
                 font=F["small"], fg=C["t3"], bg=C["card"]).pack()
    else:
        # Column header
        hdr = tk.Frame(dcard, bg=C["card2"]); hdr.pack(fill="x", pady=(0,4))
        hi  = tk.Frame(hdr,   bg=C["card2"]); hi.pack(fill="x", padx=10, pady=6)
        tk.Label(hi, text="USERNAME",    font=F["micro"], fg=C["t3"],
                 bg=C["card2"], width=16, anchor="w").pack(side="left")
        tk.Label(hi, text="ENROLLMENT",  font=F["micro"], fg=C["t3"],
                 bg=C["card2"], width=14, anchor="w").pack(side="left")
        tk.Label(hi, text="ACTION",      font=F["micro"], fg=C["t3"],
                 bg=C["card2"]).pack(side="right")

        for sname, senroll in students:
            enroll_display = senroll if senroll else "—"

            row = tk.Frame(dcard, bg=C["border"], padx=1, pady=1)
            row.pack(fill="x", pady=3)
            ri  = tk.Frame(row, bg=C["card"]);  ri.pack(fill="x")
            rii = tk.Frame(ri,  bg=C["card"]); rii.pack(fill="x", padx=10, pady=9)

            left = tk.Frame(rii, bg=C["card"]); left.pack(side="left", fill="x", expand=True)
            nr   = tk.Frame(left, bg=C["card"]); nr.pack(anchor="w")
            tk.Label(nr, text="●", font=F["micro"],
                     fg=C["accent"], bg=C["card"]).pack(side="left")
            tk.Label(nr, text=f"  {sname}", font=F["h3"],
                     fg=C["t1"], bg=C["card"]).pack(side="left")
            tk.Label(left, text=f"  Enrollment: {enroll_display}",
                     font=F["small"], fg=C["t3"], bg=C["card"]).pack(anchor="w")

            def make_del_btn(n=sname):
                return AnimBtn(rii, "Delete", lambda n=n: delete_student(n), "danger", "✕")

            make_del_btn().pack(side="right")

    # ── Pending Admin Requests ────────────────────────────
    pending = cur.execute(
        "SELECT username FROM users WHERE role='pending_admin'"
    ).fetchall()

    n_pending = len(pending)
    pcard = Card(body, gap=12)

    ph = tk.Frame(pcard, bg=C["card"]); ph.pack(fill="x", pady=(0, 6))
    Lbl(ph, "Pending Admin Requests", F["h2"], C["t1"]).pack(side="left", anchor="w")
    if n_pending:
        badge = tk.Label(ph, text=f"  {n_pending}  ", font=F["micro"],
                         fg=C["bg"], bg=C["amber"], padx=4, pady=2)
        badge.pack(side="left", padx=8)

    Lbl(pcard, "Approve or reject new admin (teacher) account requests",
        F["small"], C["t2"]).pack(anchor="w", pady=(0, 2))
    Divider(pcard, pady=10)

    if not pending:
        ef2 = tk.Frame(pcard, bg=C["card"]); ef2.pack(fill="x", pady=12)
        tk.Label(ef2, text="◈", font=("Helvetica", 24), fg=C["t4"], bg=C["card"]).pack()
        tk.Label(ef2, text="No pending requests",
                 font=F["body"], fg=C["t2"], bg=C["card"]).pack(pady=(4, 0))
        tk.Label(ef2, text="New admin requests will appear here for approval",
                 font=F["small"], fg=C["t3"], bg=C["card"]).pack()
    else:
        hdr2 = tk.Frame(pcard, bg=C["card2"]); hdr2.pack(fill="x", pady=(0, 4))
        hi2  = tk.Frame(hdr2,  bg=C["card2"]); hi2.pack(fill="x", padx=10, pady=6)
        tk.Label(hi2, text="USERNAME", font=F["micro"], fg=C["t3"],
                 bg=C["card2"], width=20, anchor="w").pack(side="left")
        tk.Label(hi2, text="ACTIONS",  font=F["micro"], fg=C["t3"],
                 bg=C["card2"]).pack(side="right")

        for (pname,) in pending:
            prow = tk.Frame(pcard, bg=C["border"], padx=1, pady=1)
            prow.pack(fill="x", pady=3)
            pri  = tk.Frame(prow, bg=C["amber_dim"]); pri.pack(fill="x")
            prii = tk.Frame(pri,  bg=C["amber_dim"]); prii.pack(fill="x", padx=10, pady=9)

            pleft = tk.Frame(prii, bg=C["amber_dim"]); pleft.pack(side="left", fill="x", expand=True)
            nr2   = tk.Frame(pleft, bg=C["amber_dim"]); nr2.pack(anchor="w")
            tk.Label(nr2, text="●", font=F["micro"],
                     fg=C["amber"], bg=C["amber_dim"]).pack(side="left")
            tk.Label(nr2, text=f"  {pname}", font=F["h3"],
                     fg=C["t1"], bg=C["amber_dim"]).pack(side="left")
            tk.Label(pleft, text="  Role: Admin  ·  Awaiting approval",
                     font=F["small"], fg=C["t3"], bg=C["amber_dim"]).pack(anchor="w")

            def make_approve_btn(n=pname):
                def approve():
                    cur.execute("UPDATE users SET role='admin' WHERE username=?", (n,))
                    conn.commit()
                    messagebox.showinfo("Approved", f"'{n}' is now an active Admin.")
                    head_admin_dash()
                return AnimBtn(prii, "Approve", approve, "success", "✓")

            def make_reject_btn(n=pname):
                def reject():
                    if messagebox.askyesno("Reject Request",
                                           f"Reject and delete '{n}'?\nThis cannot be undone."):
                        cur.execute("DELETE FROM users WHERE username=?", (n,))
                        conn.commit()
                        messagebox.showinfo("Rejected", f"'{n}' has been removed.")
                        head_admin_dash()
                return AnimBtn(prii, "Reject", reject, "danger", "✕")

            btn_row = tk.Frame(prii, bg=C["amber_dim"]); btn_row.pack(side="right")
            make_approve_btn().pack(side="left", padx=(0, 6))
            make_reject_btn().pack(side="left")

    # ── Stats ─────────────────────────────────────────────
    n_students = cur.execute("SELECT COUNT(*) FROM users WHERE role='student'").fetchone()[0]
    n_records  = cur.execute("SELECT COUNT(*) FROM attendance").fetchone()[0]
    n_admins   = cur.execute("SELECT COUNT(*) FROM users WHERE role='admin'").fetchone()[0]

    stc = Card(body, gap=10)
    Lbl(stc, "System Overview", F["micro"], C["t3"]).pack(anchor="w", pady=(0,10))
    sr = tk.Frame(stc, bg=C["card"]); sr.pack(fill="x")
    for val,label,col in [
        (n_students,"Students",  C["accent"]),
        (n_admins,  "Teachers",  C["green"]),
        (n_records, "Records",   C["purple"]),
    ]:
        sf=tk.Frame(sr,bg=C["card2"],padx=1,pady=1)
        sf.pack(side="left",fill="x",expand=True,padx=(0,6))
        si=tk.Frame(sf,bg=C["card2"]); si.pack(fill="x",padx=12,pady=10)
        tk.Label(si,text=str(val), font=F["stat"],  fg=col,     bg=C["card2"]).pack()
        tk.Label(si,text=label,    font=F["micro"], fg=C["t3"], bg=C["card2"]).pack()

    Divider(stc, pady=10)
    AnimBtn(stc,"Sign Out",login_screen,"ghost","←",full_width=True).pack(fill="x")

    tk.Frame(body, height=20, bg=C["bg"]).pack()
    StatusBar(f"Students: {n_students}  ·  Teachers: {n_admins}  ·  Pending: {n_pending}  ·  Records: {n_records}","info")


# ══════════════════════════════════════════════════════════
#  LAUNCH
# ══════════════════════════════════════════════════════════

root.bind_all("<Return>", lambda e: e.widget.invoke()
              if isinstance(e.widget, tk.Button) else None)

login_screen()
root.mainloop()